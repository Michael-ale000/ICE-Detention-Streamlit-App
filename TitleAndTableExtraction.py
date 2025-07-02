import pandas as pd
import numpy as np
import re
import os
from openpyxl import load_workbook
from sklearn.metrics.pairwise import cosine_similarity
from openai import OpenAI
from datetime import datetime


def clean_filename(name):
    return re.sub(r'[:\\/?*\[\]]', '', name).strip()[:200]

def get_embedding(text, model="text-embedding-3-small",client=None):
     # Initialize OpenAI client with API key
    response = client.embeddings.create(input=text, model=model)
    return response.data[0].embedding

def extract_tables_from_titles(filepath, sheet_name, min_length=20, suffix="FY25", skip_rows=4):
    wb = load_workbook(filepath, data_only=True)
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows())
    keywords = ["USCIS", "STATELESS", "Aliens", "Noncitizens", "Processing Disposition", "ICE"]
    noise_titles = ["ICE Release Fiscal Year"]
    titles_info = []

    for row_idx in range(skip_rows, len(rows)):
        row = rows[row_idx]
        for cell in row:
            if isinstance(cell.value, str):
                text = cell.value.strip()
                if (
                    len(text) >= min_length and
                    (text.upper().endswith(suffix.upper()) or any(k in text.upper() for k in keywords)) and
                    cell.font.bold and
                    text.upper() not in [t.upper() for t in noise_titles]
                ):
                    titles_info.append((text, row_idx, cell.column - 1))

    tables = {}
    for title, start_row, start_col in titles_info:
        max_row = start_row + 1
        max_col = start_col

        while max_row < len(rows):
            if all(cell.value is None for cell in rows[max_row][start_col:start_col + 20]):
                break
            max_row += 1

        while True:
            col_empty = True
            for r in range(start_row + 1, max_row):
                if max_col >= len(rows[r]):
                    continue
                if rows[r][max_col].value is not None:
                    col_empty = False
                    break
            if col_empty:
                break
            max_col += 1

        table = []
        for r in range(start_row + 1, max_row):
            row_data = [rows[r][c].value if c < len(rows[r]) else None for c in range(start_col, max_col)]
            table.append(row_data)

        tables[title] = table

    return tables

def process_footnotes(filepath):
    df = pd.read_excel(filepath, sheet_name='Footnotes', skiprows=3)
    df['Term'] = df['Term'].ffill()
    df_clean = df.groupby('Term', as_index=False).agg({'Definition': ' '.join})
    df_clean['EID_Date'] = df_clean['Definition'].str.extract(r'EID as of (\d{2}/\d{2}/\d{4})')
    return df_clean[df_clean['EID_Date'].notnull()]

def Extraction_title_and_data(filepath, source_filename=None,api_key=None, save_excel=True,):
    client = OpenAI(api_key=api_key) 
    # Determine the filename to use for pattern matching
    filename_for_regex = source_filename if source_filename else filepath

    fy_match = re.search(r'FY(\d{2})', filename_for_regex, re.IGNORECASE)
    if fy_match:
        fy_short = fy_match.group(1)
        fy_full = f"FY20{fy_short}"
        fiscal_year = f"FY{fy_short}"
    else:
        fy_short = "XX"
        fy_full = "FY20XX"
        fiscal_year = "FYXX"

    sheet_name = f"Detention {fiscal_year}"
    suffix = fy_full

    matches = re.findall(r'\d{8}', filename_for_regex)
    if matches:
        date_str = matches[-1]
        release_date = datetime.strptime(date_str, '%m%d%Y').strftime('%Y-%m-%d')
    else:
        release_date = None

    output_dir = os.path.splitext(filepath)[0]
    os.makedirs(output_dir, exist_ok=True)

    tables_info = extract_tables_from_titles(filepath, sheet_name=sheet_name, suffix=suffix)
    footnote_df = process_footnotes(filepath)

    footnote_terms = footnote_df["Term"].tolist()
    footnote_eids = footnote_df["EID_Date"].tolist()

    title_embeddings = [get_embedding(title,client=client) for title in tables_info.keys()]
    footnote_embeddings = [get_embedding(term,client=client) for term in footnote_terms]

    similarity_matrix = cosine_similarity(title_embeddings, footnote_embeddings)

    structured_tables = {}
    source_filename_basename = os.path.basename(filename_for_regex)

    for i, title in enumerate(tables_info.keys(), start=1):
        best_idx = np.argmax(similarity_matrix[i - 1])
        best_score = similarity_matrix[i - 1][best_idx]
        matched_eid = footnote_eids[best_idx] if best_score > 0.5 else None

        table_data = tables_info[title]
        if not table_data:
            continue

        new_columns = ["EID", "Source_filename", "Release_date", "Table_name", "Table_code"]
        table_data[0].extend(new_columns)

        for row in table_data[1:]:
            row.extend([matched_eid, source_filename_basename, release_date, title, i])

        structured_tables[title] = {
            "title": title,
            "EID": matched_eid,
            "table_data": table_data
        }

    table_dataframes = {}
    for title, info in structured_tables.items():
        data_rows = info["table_data"]
        if not data_rows:
            continue

        df = pd.DataFrame(data_rows)
        new_columns = [str(col) if col is not None else "" for col in df.iloc[0]]
        df.columns = new_columns
        df = df[1:].dropna(how='all')

        table_dataframes[title] = df

    if save_excel and table_dataframes:
        output_path = os.path.join(output_dir, clean_filename(source_filename_basename.replace('.xlsx', '_processed.xlsx')))
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            for i, (title, df) in enumerate(table_dataframes.items(), start=1):
                sheet_name = f"Table {i}"
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        print(f"Saved Excel file to: {output_path}")

    return table_dataframes
